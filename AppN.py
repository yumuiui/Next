import re
import zipfile
import tempfile
from datetime import datetime, date
from io import BytesIO
from pathlib import Path

import pandas as pd
import pdfplumber
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Next Supply | Separador de OPS", page_icon="N", layout="wide")

st.markdown("""<style>
:root{
    --bg:#02053d;--panel:#060a53;--panel2:#08106a;
    --border:rgba(63,93,255,0.35);--border-soft:rgba(255,255,255,0.08);
    --text:#f5f7ff;--muted:#a9b4ea;
    --primary:#1237ff;--primary2:#3f5dff;
    --shadow:0 18px 60px rgba(0,0,0,0.28);
    --radius-xl:28px;--radius-lg:22px;
}
.stApp{
    background:
        radial-gradient(circle at top left,rgba(63,93,255,0.20),transparent 28%),
        radial-gradient(circle at top right,rgba(18,55,255,0.16),transparent 24%),
        linear-gradient(180deg,#010332 0%,var(--bg) 100%);
    color:var(--text);
}
.block-container{max-width:1400px;padding-top:1.4rem;padding-bottom:2.5rem;}
[data-testid="stHeader"]{background:transparent;}
[data-testid="stSidebar"]{background:var(--panel) !important;border-right:1px solid var(--border-soft) !important;}
[data-testid="stSidebar"] *{color:var(--text) !important;}
h1,h2,h3,h4,h5,h6,p,label,div,span{color:var(--text);}
.topbar{display:flex;align-items:center;justify-content:space-between;gap:24px;background:rgba(4,8,70,0.80);border:1px solid var(--border-soft);border-radius:24px;padding:18px 24px;box-shadow:var(--shadow);margin-bottom:22px;}
.brand-wrap{display:flex;align-items:center;gap:16px;}
.brand-mark{width:58px;height:58px;border-radius:14px;background:linear-gradient(135deg,var(--primary2),var(--primary));display:flex;align-items:center;justify-content:center;color:white !important;font-size:1.8rem;font-weight:800;box-shadow:0 10px 30px rgba(18,55,255,0.35);}
.brand-title{font-size:1.5rem;font-weight:800;line-height:1;margin:0;}
.brand-subtitle{margin:4px 0 0 0;color:var(--muted) !important;font-size:0.95rem;}
.hero{background:linear-gradient(180deg,rgba(6,10,83,0.92) 0%,rgba(4,7,58,0.92) 100%);border:1px solid var(--border);border-radius:var(--radius-xl);padding:30px;box-shadow:var(--shadow);margin-bottom:20px;}
.hero-title{font-size:2.2rem;font-weight:800;margin-bottom:0.4rem;}
.hero-sub{color:var(--muted) !important;font-size:1rem;margin-bottom:1.4rem;max-width:900px;}
.metrics{display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:16px;}
.mcard{background:rgba(255,255,255,0.04);border:1px solid var(--border-soft);border-radius:var(--radius-lg);padding:18px 20px;backdrop-filter:blur(8px);}
.mcard-label{color:var(--muted) !important;font-size:0.9rem;margin-bottom:8px;}
.mcard-val{font-size:1.7rem;font-weight:800;line-height:1.1;color:var(--text) !important;}
.mcard-val.accent{color:#7b9fff !important;}
.sec-title{font-size:1rem;font-weight:700;border-left:3px solid var(--primary2);padding-left:12px;margin:20px 0 12px 0;}
.alerta-grid{display:grid;grid-template-columns:repeat(3,1fr);gap:14px;margin-bottom:4px;}
.alerta-card{border-radius:16px;padding:20px 22px;display:flex;flex-direction:column;gap:6px;}
.alerta-card.red{background:rgba(220,38,38,0.15);border:1px solid rgba(220,38,38,0.45);}
.alerta-card.yellow{background:rgba(234,179,8,0.12);border:1px solid rgba(234,179,8,0.40);}
.alerta-card.green{background:rgba(34,197,94,0.10);border:1px solid rgba(34,197,94,0.35);}
.alerta-label{font-size:0.8rem;font-weight:600;opacity:0.75;}
.alerta-val{font-size:2rem;font-weight:900;line-height:1;}
.alerta-card.red .alerta-val{color:#f87171 !important;}
.alerta-card.yellow .alerta-val{color:#fbbf24 !important;}
.alerta-card.green .alerta-val{color:#4ade80 !important;}
.alerta-sub{font-size:0.72rem;opacity:0.6;}
div[data-testid="stFileUploader"] section{background:rgba(255,255,255,0.03) !important;border:1px dashed rgba(88,117,255,0.45) !important;border-radius:18px !important;padding:8px !important;}
div[data-testid="stDataFrame"]{border:1px solid var(--border-soft) !important;border-radius:18px !important;overflow:hidden !important;background:rgba(255,255,255,0.02) !important;}
.stDownloadButton>button,.stButton>button{border-radius:14px !important;min-height:46px !important;padding:0.72rem 1.15rem !important;font-weight:800 !important;border:1px solid transparent !important;color:white !important;background:linear-gradient(135deg,var(--primary),var(--primary2)) !important;box-shadow:0 10px 24px rgba(18,55,255,0.28) !important;}
.stAlert{border-radius:16px !important;}
.hr{height:1px;background:linear-gradient(90deg,var(--primary2),rgba(63,93,255,0.1),transparent);border:none;margin:24px 0;}
@media(max-width:1100px){.metrics{grid-template-columns:repeat(2,minmax(0,1fr));}.alerta-grid{grid-template-columns:1fr;}}
@media(max-width:700px){.metrics{grid-template-columns:1fr;}.hero-title{font-size:1.8rem;}}
</style>""", unsafe_allow_html=True)


# ── Utilitarios ───────────────────────────────────────────────────────────────

def clean(text):
    text = (text or "").replace("\u00ad", "")
    text = re.sub(r"Resumo extra[i\u00ed]do por[^\n]*(NEXTSUPPLY\d+\))[^\n]*", "", text, flags=re.I)
    text = re.sub(r"\bP[a\u00e1]g[.:]\s*\d+/\d+\b", "", text, flags=re.I)
    text = re.sub(r"(?is)Resumo extra[i\u00ed]do por.*?(?=\n|$)", "", text)
    # Remove cabecalho de pagina do Petronect que aparece no meio dos blocos
    text = re.sub(
        r"(?im)^Resumo da Oportunidade\s+N[u\u00fa]mero da Oportunidade\s*\n.*?\d{10}\s*\n?",
        "",
        text,
    )
    text = re.sub(
        r"(?im)^N[u\u00fa]mero da Oportunidade\s*\n\s*\d{10}\s*\n\s*Resumo da Oportunidade\s*\n[^\n]+\n?",
        "",
        text,
    )
    text = re.sub(r"(?im)^\s*Resumo da Oportunidade\s*$", "", text)
    text = re.sub(r"(?is)ENDERE[C\u00c7]O DE ENTREGA E FATURAMENTO:.*?(?=\nDados do Item|\nDescri|\nDeclara|\Z)", "", text)
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\s*\n\s*", "\n", text)
    return text.strip()


def one_line(text):
    text = clean(text).replace("\n", " ")
    text = re.sub(r"\(NEXTSUPPLY\d+\)", "", text, flags=re.I)
    return re.sub(r"\s{2,}", " ", text).strip()


def brand_regex(term):
    parts = re.split(r"\s+", str(term).strip().lower())
    return r"\b" + r"[\s\-]?".join(map(re.escape, parts)) + r"\b"


# ── Categorias ────────────────────────────────────────────────────────────────

CATEGORIAS = [
    ("Valvula",           ["valv", "ball valve", "gate valve", "check valve", "globo", "borboleta", "esfera", "gaveta", "retencao", "alivio"]),
    ("Tubo",              ["tubo ", "tubulac", " piping", "conduc", " pipe "]),
    ("Junta / Vedacao",   ["junta", "gaxeta", "oring", "o-ring", "gasket", "vedac", "anel de veda", "anel lip"]),
    ("Rele",              ["rele ", "relay", "rele de protecao", "rele de tensao", "rele aux"]),
    ("Bomba",             ["bomba", " pump ", "booster"]),
    ("Filtro",            ["filtro", "filter", "coador", "strainer"]),
    ("Instrumento",       ["instrumento", "sensor", "transmissor", "medidor", "manometro", "pressostato", "termometro"]),
    ("Cabo / Fio",        ["cabo ", "cable", "condutor", "teldor"]),
    ("Flange",            ["flange"]),
    ("Conector / Fitting",["fitting", "niple", "cotovelo", " tee "]),
    ("Parafuso / Fixador",["parafuso", "arruela", " bolt", " stud "]),
    ("Mangueira",         ["mangueira", " hose "]),
    ("Conjunto / Kit",    ["conj.", "conjunto especif", " kit ", "assembly"]),
    ("Relatorio / Doc",   ["manual de operac", "manual tecnico", "documento tecnico", "relatorio tecnico"]),
]


def categorize(text):
    t = (text or "").lower()
    for cat, keywords in CATEGORIAS:
        for kw in keywords:
            if kw in t:
                return cat
    return "Outros"


# ── Deteccao de recorrencia ───────────────────────────────────────────────────

def detect_recurring(df_new, df_hist):
    if df_hist.empty or df_new.empty:
        return pd.Series([False] * len(df_new), index=df_new.index)

    def norm(s):
        return re.sub(r"\s+", " ", str(s or "").lower().strip())

    hist_fab  = set(df_hist["Fabricante/PN"].dropna().map(norm)) - {"", "nan"}
    hist_desc = set(df_hist["Descricao longa do item"].dropna().map(norm)) - {"", "nan"}

    def is_rec(row):
        fab  = norm(row.get("Fabricante/PN", ""))
        desc = norm(row.get("Descricao longa do item", ""))
        return (fab and fab in hist_fab) or (desc and desc in hist_desc)

    return df_new.apply(is_rec, axis=1)


# ── Alertas de prazo ─────────────────────────────────────────────────────────

def classify_prazo(data_str):
    try:
        d = datetime.strptime(str(data_str).strip(), "%d/%m/%Y").date()
        delta = (d - date.today()).days
        if delta <= 1:   return delta, "red"
        elif delta <= 3: return delta, "yellow"
        else:            return delta, "green"
    except Exception:
        return None, "green"


def render_alertas(df):
    if df is None or df.empty or "Data (cotacao)" not in df.columns:
        return
    ops = df.drop_duplicates(subset=["Numero da Oportunidade"])[["Numero da Oportunidade", "Data (cotacao)"]].copy()
    ops["_dias"], ops["_cor"] = zip(*ops["Data (cotacao)"].map(classify_prazo))
    n_red    = int((ops["_cor"] == "red").sum())
    n_yellow = int((ops["_cor"] == "yellow").sum())
    n_green  = int((ops["_cor"] == "green").sum())

    sec("Alertas de Prazo")
    st.markdown(
        '<div class="alerta-grid">'
        '<div class="alerta-card red"><div class="alerta-label">URGENTE - Vence hoje / amanha</div>'
        '<div class="alerta-val">' + str(n_red) + '</div>'
        '<div class="alerta-sub">' + (str(n_red) + " oportunidade(s)" if n_red else "Nenhuma urgente") + '</div></div>'
        '<div class="alerta-card yellow"><div class="alerta-label">ATENCAO - Vence em 2 a 3 dias</div>'
        '<div class="alerta-val">' + str(n_yellow) + '</div>'
        '<div class="alerta-sub">' + (str(n_yellow) + " oportunidade(s)" if n_yellow else "Nenhuma em atencao") + '</div></div>'
        '<div class="alerta-card green"><div class="alerta-label">OK - Prazo confortavel</div>'
        '<div class="alerta-val">' + str(n_green) + '</div>'
        '<div class="alerta-sub">' + (str(n_green) + " oportunidade(s)" if n_green else "Nenhuma") + '</div></div>'
        '</div>',
        unsafe_allow_html=True
    )
    if n_red > 0:
        red_nums = ops[ops["_cor"] == "red"]["Numero da Oportunidade"]
        with st.expander("Ver oportunidades urgentes"):
            urgentes = df[df["Numero da Oportunidade"].isin(red_nums)]
            show_cols = [c for c in ["Numero da Oportunidade", "Data (cotacao)", "Responsavel", "Descricao longa do item"] if c in urgentes.columns]
            st.dataframe(urgentes[show_cols], use_container_width=True, hide_index=True)


# ── Extracao PDF ──────────────────────────────────────────────────────────────

def extract_header(raw, fallback=""):
    m = re.search(r"N[u\u00fa]mero da Oportunidade\s*(\d{10})", raw, flags=re.S | re.I)
    if m:
        numero = m.group(1)
    else:
        fb = re.search(r"(\d{10})", str(fallback))
        numero = fb.group(1) if fb else ""

    def get(pattern):
        r = re.search(pattern, raw, flags=re.I)
        return r.group(1).strip() if r else ""

    tipo = get(r"Tipo de Oportunidade\s*([^\n]+)")
    crit = get(r"Crit[e\u00e9]rio de Julgamento\s*([^\n]+)")
    fim_raw = get(r"Fim do per[i\u00ed]odo de cota[c\u00e7][a\u00e3]o\s*([0-9]{2}\.[0-9]{2}\.[0-9]{4}\s*/\s*[0-9]{2}:[0-9]{2}:[0-9]{2})")
    fim = re.sub(r"\s*/\s*", " / ", fim_raw)
    ml = re.search(r"Local de Entrega\s*(.*?)\nInforma[c\u00e7][o\u00f5]es do Comprador", raw, flags=re.S | re.I)
    local = "; ".join(x.strip() for x in ml.group(1).split("\n") if x.strip()) if ml else ""
    return {"numero": numero, "tipo": tipo, "crit": crit, "fim": fim, "local": local}


def extract_item_id(block):
    for pat, flags in [
        (r"(?m)^\s*(\d{1,6})\s+\S", 0),
        (r"(?i)N[u\u00fa]mero\s+Descri.*?\n(\d+)\s", re.S),
        (r"(?i)N[u\u00fa]mero\s*\n\s*do item\s*(?:\n|\s)+([0-9A-Za-z\.\-]+)", 0),
    ]:
        m = re.search(pat, block, flags=flags)
        if m:
            return re.sub(r"\D", "", m.group(1))
    return ""


def extract_qty_unit(block, item_id):
    if not item_id:
        return "", ""
    patterns = [
        rf"(?im)^\s*{re.escape(item_id)}\s+.*?\bMaterial\b\s+([0-9]+(?:\.[0-9]{{3}})*(?:,[0-9]+)?)\s+([A-Za-z\u00c0-\u00ff]+)\s+\d{{2}}\.\d{{2}}\.\d{{4}}\b",
        rf"(?im)^\s*{re.escape(item_id)}\s+.*?\b([0-9]+(?:\.[0-9]{{3}})*(?:,[0-9]+)?)\s+([A-Za-z\u00c0-\u00ff]+)\s+\d{{2}}\.\d{{2}}\.\d{{4}}\b",
        rf"(?is)\b{re.escape(item_id)}\b.*?\bMaterial\b\s+([0-9]+(?:\.[0-9]{{3}})*(?:,[0-9]+)?)\s+([A-Za-z\u00c0-\u00ff]+)\b",
        r"(?is)\bQuantidade\b\s*([0-9]+(?:\.[0-9]{3})*(?:,[0-9]+)?)\s+\b([A-Za-z\u00c0-\u00ff]+)\b",
    ]
    for p in patterns:
        m = re.search(p, block)
        if m:
            qty = m.group(1).strip()
            qty_clean = qty.replace(".", "").replace(",", ".")
            try:
                qty_num = float(qty_clean)
                qty_fmt = str(int(qty_num)) if qty_num == int(qty_num) else str(qty_num)
            except Exception:
                qty_fmt = qty
            return qty_fmt, m.group(2).strip()
    return "", ""


def extract_desc(block):
    b = re.split(r"(?i)ENDERE[C\u00c7]O DE ENTREGA", block, maxsplit=1)[0]
    m = re.search(
        r"(?is)Descri[c\u00e7][a\u00e3]o de Item\s*(.*?)(?:\nDescri[c\u00e7][a\u00e3]o longa|\Z)",
        b
    )
    result = one_line(m.group(1)) if m else ""
    return result if result.strip() else "Sem descricao"


def extract_long(block):
    b = re.split(r"(?i)ENDERE[C\u00c7]O DE ENTREGA", block, maxsplit=1)[0]
    b = re.split(r"(?i)Declara[c\u00e7][o\u00f5]es envolvidas", b, maxsplit=1)[0]
    b = re.split(r"(?i)Resumo extra[i\u00ed]do por", b, maxsplit=1)[0]
    b = re.split(r"(?i)Resumo da Oportunidade", b, maxsplit=1)[0]
    m = re.search(
        r"(?is)Descri[c\u00e7][a\u00e3]o longa do item\s*(.*?)$|Descri[c\u00e7][a\u00e3]o longa\s*(.*?)$",
        b
    )
    if not m:
        return ""
    result = one_line(m.group(1) or m.group(2))
    result = re.sub(r"\s*\bP[a\u00e1]g:\s*\d+/\d+.*$", "", result, flags=re.I).strip()
    result = re.sub(r"\s*\b\d{2}\.\d{2}\.\d{4}\s*(?:[a\u00e0]s)?\s*\d{2}:\d{2}:\d{2}.*$", "", result, flags=re.I).strip()
    result = re.sub(r"\s*\(NEXTSUPPLY\d+\).*$", "", result, flags=re.I).strip()
    result = re.sub(r"[\s;,]+$", "", result).strip()
    return result


def extract_manufacturer(block):
    b = re.split(r"(?i)ENDERE[C\u00c7]O DE ENTREGA", block, maxsplit=1)[0]
    b = re.split(r"(?i)Resumo extra[i\u00ed]do por", b, maxsplit=1)[0]
    b = re.split(r"(?i)Resumo da Oportunidade", b, maxsplit=1)[0]
    m = re.search(r"(?i)Tp:\s*(.+?)(?=\s*-{5,}|\s*ENDERE|\s*Resumo|\s*Dados|\s*Declara|$)", b, re.S)
    if m:
        fab = one_line(m.group(1))
        fab = re.sub(r"[\s/\-|:;,]+$", "", fab).strip()
        if fab:
            return fab
    m2 = re.search(r"(?i)FABRICANTE:\s*(\S+)", b)
    if m2:
        return m2.group(1).strip()
    m3 = re.search(r"(?i)REFER[E\u00ca]NCIA:\s*([^\n/]+)", b)
    if m3:
        return m3.group(1).strip()
    return ""


# ── Atribuicao ────────────────────────────────────────────────────────────────

DEFAULT_TEAM = [
    {"name": "Viviana", "brands": ["kongsberg", "yamada", "dnh", "evac", "steyr"]},
    {"name": "Helio",   "brands": ["abb", "schneider", "siemens", "rittal", "phoenix", "weidmuller", "rockwell"]},
    {"name": "Mayara",  "brands": ["skf", "emerson"]},
]


def sidebar_team():
    with st.sidebar:
        st.markdown("## Equipe")
        st.caption("Configure os responsaveis e as marcas de cada um.")
        n = st.number_input("Nr de responsaveis", min_value=1, max_value=10, value=len(DEFAULT_TEAM), step=1)
        team = []
        for i in range(int(n)):
            default = DEFAULT_TEAM[i] if i < len(DEFAULT_TEAM) else {"name": "Responsavel " + str(i+1), "brands": []}
            with st.expander("Responsavel " + str(i+1), expanded=(i == 0)):
                name = st.text_input("Nome", value=default["name"], key="rname_" + str(i))
                brands_raw = st.text_area(
                    "Marcas associadas (uma por linha)",
                    value="\n".join(default["brands"]),
                    height=100,
                    key="rbrands_" + str(i),
                )
                brands = [b.strip().lower() for b in brands_raw.splitlines() if b.strip()]
                team.append({"name": name, "brands": brands})
        st.markdown("---")
        st.markdown("Next Supply\nSeparador de OPS Petronect")
    return team


def assign(df, team):
    if df.empty:
        df["Responsavel"] = pd.Series(dtype="object")
        return df

    corpus = (
        df["Descricao de Item"].fillna("") + " " +
        df["Descricao longa do item"].fillna("") + " " +
        df["Fabricante/PN"].fillna("")
    ).str.lower()

    df["Responsavel"] = pd.NA

    # Passo 1 — marcas fixas por responsavel
    for member in team:
        for brand in member.get("brands", []):
            mask = df["Responsavel"].isna() & corpus.str.contains(brand_regex(brand), regex=True, na=False)
            df.loc[mask, "Responsavel"] = member["name"]

    # Passo 2 — balanceamento proporcional dos restantes
    names = [m["name"] for m in team]
    if not names:
        return df

    total_na = int(df["Responsavel"].isna().sum())
    if total_na == 0:
        return df

    n = len(names)
    base = total_na // n
    extra = total_na % n
    targets = {names[i]: base + (1 if i < extra else 0) for i in range(n)}
    counts = {name: 0 for name in names}

    chave = (
        df["Descricao de Item"].fillna("").str.strip().str.lower() + "|" +
        df["Descricao longa do item"].fillna("").str.strip().str.lower() + "|" +
        df["Fabricante/PN"].fillna("").str.strip().str.lower()
    )

    for _, grp in df[df["Responsavel"].isna()].groupby(chave, sort=False):
        chosen = max(names, key=lambda name: targets[name] - counts[name])
        df.loc[grp.index, "Responsavel"] = chosen
        counts[chosen] += len(grp)

    return df



# ── Historico de Precos por PN ────────────────────────────────────────────────

def load_historico(file_bytes):
    """Le a aba DADOS da planilha de analise e retorna DataFrame limpo."""
    try:
        df = pd.read_excel(BytesIO(file_bytes), sheet_name="DADOS", dtype=str)
        df.columns = [c.strip() for c in df.columns]
        # Normaliza fabricante e PN
        df["_fab"] = df["FABRICANTE"].fillna("").str.strip().str.upper()
        df["_pn"]  = df["PART NUMBER"].fillna("").str.strip().str.upper()
        return df.reset_index(drop=True)
    except Exception:
        return pd.DataFrame()


def enriquecer_com_historico(df_extr, df_preco):
    """
    Para cada item extraido, busca historico assim:
      1. Filtra historico pelo FABRICANTE do item
      2. Para cada PN analisado desse fabricante, faz ctrl-F na descricao longa
      3. Se o PN aparecer em qualquer parte da descricao -> match encontrado
      4. Tambem faz ctrl-F com os nomes de fabricante do historico na descricao
      5. Agrega os matches e retorna ultima analise, dif ultima, dif media, OP e data

    Logica: DIFERENCA negativa = ganhamos. DIFERENCA positiva = perdemos.
    """
    cols_out = [
        "Hist: Ultima analise", "Hist: Dif ultima",
        "Hist: Dif media hist", "Hist: OP referencia", "Hist: Data analise",
    ]
    for c in cols_out:
        df_extr[c] = ""

    if df_preco.empty:
        return df_extr

    # Identificar colunas chave uma vez
    col_dif = next((c for c in df_preco.columns if "DIFE" in c.upper() and "VALOR" in c.upper()), None)
    col_pct = next((c for c in df_preco.columns if c.strip() == "%"), None)
    col_res = next((c for c in df_preco.columns if "RESULTADO" in c.upper()), None)
    col_dat = next((c for c in df_preco.columns if c.strip() == "DATA"), None)
    col_op  = next((c for c in df_preco.columns if c.strip() == "OP"), None)

    if not col_dif:
        return df_extr

    # Pre-calcular fabricantes e PNs do historico uma vez (performance)
    fabs_hist = [(fab, df_preco[df_preco["_fab"] == fab]) for fab in df_preco["_fab"].dropna().unique() if len(fab) >= 4]
    pns_hist_todos = [p for p in df_preco["_pn"].dropna().unique() if len(p) >= 6]

    def fmt_brl(v):
        sinal = "+" if v > 0 else ""
        return sinal + "R$ {:,.2f}".format(v).replace(",", "X").replace(".", ",").replace("X", ".")

    def fmt_pct(v):
        sinal = "+" if v > 0 else ""
        return sinal + "{:.1f}%".format(v * 100)

    def norm(s):
        return str(s or "").strip().upper()

    for idx, row in df_extr.iterrows():
        fab_item  = norm(row.get("Fabricante/PN", ""))
        desc_longa = norm(row.get("Descricao longa do item", ""))

        if not fab_item and not desc_longa:
            df_extr.at[idx, "Hist: Ultima analise"] = "Sem historico"
            continue

        matches = pd.DataFrame()

        # --- Passo 1: filtrar historico pelo fabricante extraido (match direto) ---
        if fab_item:
            mask1 = df_preco["_fab"].apply(lambda h: bool(h and h in fab_item))
            mask2 = df_preco["_fab"].apply(lambda h: bool(h and fab_item in h))
            sub_fab = df_preco[mask1 | mask2].copy()

            # --- Passo 2: ctrl-F — busca cada PN desse fabricante na descricao ---
            if not sub_fab.empty and desc_longa:
                pns_fab = [p for p in sub_fab["_pn"].dropna().unique() if len(p) >= 4]
                matched_rows = [sub_fab[sub_fab["_pn"] == pn] for pn in pns_fab if pn in desc_longa]
                if matched_rows:
                    matches = pd.concat(matched_rows, ignore_index=True)

        # --- Passo 3: ctrl-F com FABRICANTE do historico na descricao longa ---
        # (funciona quando o fabricante nao foi extraido do PDF mas aparece na descricao)
        if matches.empty and desc_longa:
            matched_rows = []
            for fab_h, sub_h in fabs_hist:
                if fab_h in desc_longa:  # ctrl-F: nome do fabricante aparece na descricao
                    # Dentro desse fabricante, busca tambem os PNs na descricao
                    pns_h = [p for p in sub_h["_pn"].dropna().unique() if len(p) >= 4]
                    pn_rows = [sub_h[sub_h["_pn"] == pn] for pn in pns_h if pn in desc_longa]
                    if pn_rows:
                        matched_rows.extend(pn_rows)
                    else:
                        # Fabricante achou mas PN nao: usa todos os registros do fabricante
                        matched_rows.append(sub_h)
            if matched_rows:
                matches = pd.concat(matched_rows, ignore_index=True)

        # --- Passo 4 (fallback): ctrl-F de todos os PNs do historico na descricao ---
        if matches.empty and desc_longa:
            matched_rows = [df_preco[df_preco["_pn"] == pn] for pn in pns_hist_todos if pn in desc_longa]
            if matched_rows:
                matches = pd.concat(matched_rows, ignore_index=True)

        if matches.empty:
            df_extr.at[idx, "Hist: Ultima analise"] = "Sem historico"
            continue

        # --- Passo 5: calcular metricas dos matches ---
        matches = matches.copy()
        matches["_dif"] = pd.to_numeric(matches[col_dif], errors="coerce")
        if col_pct:
            matches["_pct"] = pd.to_numeric(matches[col_pct], errors="coerce")
        if col_dat:
            matches = matches.sort_values(col_dat, ascending=True)

        if col_res:
            com_res = matches[matches[col_res].fillna("").str.strip().str.lower().isin(["ganhamos", "perdemos"])]
        else:
            com_res = matches

        if com_res.empty:
            df_extr.at[idx, "Hist: Ultima analise"] = "Sem historico"
            continue

        ultima  = com_res.iloc[-1]
        res_ult = str(ultima.get(col_res, "") or "").strip()
        dif_ult = ultima.get("_dif", None)
        pct_ult = ultima.get("_pct", None) if col_pct else None

        df_extr.at[idx, "Hist: Ultima analise"] = res_ult if res_ult else "Sem historico"

        if pd.notna(dif_ult):
            txt = fmt_brl(dif_ult)
            if pd.notna(pct_ult):
                txt += " (" + fmt_pct(pct_ult) + ")"
            df_extr.at[idx, "Hist: Dif ultima"] = txt

        if col_pct and "_pct" in matches.columns:
            vals = matches["_pct"].dropna()
            if not vals.empty:
                df_extr.at[idx, "Hist: Dif media hist"] = fmt_pct(vals.mean()) + " medio"

        if col_op:
            op_ref = str(ultima.get(col_op, "") or "").strip()
            if op_ref:
                df_extr.at[idx, "Hist: OP referencia"] = op_ref

        if col_dat:
            data_ref = str(ultima.get(col_dat, "") or "").strip()
            if data_ref and data_ref != "nan":
                df_extr.at[idx, "Hist: Data analise"] = data_ref

    return df_extr


# ── Pipeline ──────────────────────────────────────────────────────────────────

COLS = [
    "Numero da Oportunidade", "Tipo de Oportunidade", "Criterio de Julgamento",
    "Fim do periodo de cotacao", "Data (cotacao)", "Hora (cotacao)",
    "Local de Entrega", "Item", "Quantidade", "Unidade de medida",
    "Descricao de Item", "Descricao longa do item",
    "Fabricante/PN", "Categoria", "Recorrente", "Responsavel",
    "Hist: Ultima analise", "Hist: Dif ultima", "Hist: Dif media hist",
    "Hist: OP referencia", "Hist: Data analise",
]


def process_zip(zip_bytes, team, df_hist=None, df_preco=None):
    rows, log = [], []
    with tempfile.TemporaryDirectory() as tmp:
        tmp = Path(tmp)
        with zipfile.ZipFile(BytesIO(zip_bytes)) as z:
            z.extractall(tmp)
        pdfs = sorted(tmp.rglob("*.pdf"))
        if not pdfs:
            return pd.DataFrame(columns=COLS), ["Nenhum PDF encontrado no ZIP."]
        for pdf in pdfs:
            try:
                with pdfplumber.open(str(pdf)) as p:
                    raw_pages = "\n".join((pg.extract_text() or "") for pg in p.pages)
                header = extract_header(raw_pages, pdf.stem)   # antes do clean
                raw    = clean(raw_pages)
                blocks = re.split(r"(?i)\bDados do Item\b", raw)[1:]
                if not blocks:
                    log.append("AVISO: " + pdf.name + " sem itens.")
                    continue
                count = 0
                for block in blocks:
                    item_id = extract_item_id(block)
                    if not item_id:
                        continue
                    qty, unit = extract_qty_unit(block, item_id)
                    desc   = extract_desc(block)
                    long_d = extract_long(block)
                    fab    = extract_manufacturer(block)
                    rows.append({
                        "Numero da Oportunidade":    header["numero"],
                        "Tipo de Oportunidade":      header["tipo"],
                        "Criterio de Julgamento":    header["crit"],
                        "Fim do periodo de cotacao": header["fim"],
                        "Local de Entrega":          header["local"],
                        "Item":                      item_id,
                        "Quantidade":                qty,
                        "Unidade de medida":         unit,
                        "Descricao de Item":         desc,
                        "Descricao longa do item":   long_d,
                        "Fabricante/PN":             fab,
                        "Categoria":                 categorize(long_d or desc),
                    })
                    count += 1
                log.append("OK: " + pdf.name + " - " + str(count) + " item(ns).")
            except Exception as e:
                log.append("ERRO: " + pdf.name + " - " + str(e))

    if not rows:
        return pd.DataFrame(columns=COLS), log

    df = pd.DataFrame(rows).drop_duplicates(subset=["Numero da Oportunidade", "Item"], keep="first")
    dt = pd.to_datetime(
        df["Fim do periodo de cotacao"].astype(str).str.replace(" / ", " ", regex=False),
        format="%d.%m.%Y %H:%M:%S", errors="coerce",
    )
    df["Data (cotacao)"] = dt.dt.strftime("%d/%m/%Y")
    df["Hora (cotacao)"] = dt.dt.strftime("%H:%M:%S")

    hist = df_hist if df_hist is not None and not df_hist.empty else pd.DataFrame()
    df["Recorrente"] = detect_recurring(df, hist).map({True: "Sim", False: "Nao"})

    df = assign(df, team)

    # Enriquece com historico de precos se disponivel
    if df_preco is not None and not df_preco.empty:
        df = enriquecer_com_historico(df, df_preco)

    for col in COLS:
        if col not in df.columns:
            df[col] = ""
    return df[COLS], log


# ── Excel formatado ───────────────────────────────────────────────────────────

HDR_FILL  = PatternFill("solid", fgColor="02053D")
HDR_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
ODD_FILL  = PatternFill("solid", fgColor="F0F2FF")
EVEN_FILL = PatternFill("solid", fgColor="FFFFFF")
REC_FILL  = PatternFill("solid", fgColor="FFF3CD")
BODY_FONT = Font(name="Arial", size=10)
CENTER    = Alignment(horizontal="center", vertical="center", wrap_text=False)
LEFT      = Alignment(horizontal="left", vertical="center", wrap_text=True)
THIN      = Side(style="thin", color="C8CEEA")
BORDER    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COL_LABELS = {
    "Numero da Oportunidade":    "Nr Oportunidade",
    "Tipo de Oportunidade":      "Tipo",
    "Criterio de Julgamento":    "Criterio",
    "Fim do periodo de cotacao": "Prazo",
    "Data (cotacao)":            "Data",
    "Hora (cotacao)":            "Hora",
    "Local de Entrega":          "Local",
    "Item":                      "Item",
    "Quantidade":                "Qtd",
    "Unidade de medida":         "Un",
    "Descricao de Item":         "Descricao Curta",
    "Descricao longa do item":   "Descricao Longa",
    "Fabricante/PN":             "Fabricante / PN",
    "Categoria":                 "Categoria",
    "Recorrente":                "Recorrente",
    "Responsavel":               "Responsavel",
    "Hist: Ultima analise":      "Ultima Analise",
    "Hist: Dif ultima":          "Dif Ultima (R$/%)",
    "Hist: Dif media hist":      "Dif Media Hist",
    "Hist: OP referencia":       "OP Referencia",
    "Hist: Data analise":        "Data Analise",
}

COL_WIDTHS = {
    "Numero da Oportunidade":    16,
    "Tipo de Oportunidade":      22,
    "Criterio de Julgamento":    16,
    "Fim do periodo de cotacao": 22,
    "Data (cotacao)":            12,
    "Hora (cotacao)":            10,
    "Local de Entrega":          26,
    "Item":                       6,
    "Quantidade":                 8,
    "Unidade de medida":          8,
    "Descricao de Item":         32,
    "Descricao longa do item":   50,
    "Fabricante/PN":             22,
    "Categoria":                 16,
    "Recorrente":                12,
    "Responsavel":               14,
    "Hist: Ultima analise":      14,
    "Hist: Dif ultima":          22,
    "Hist: Dif media hist":      16,
    "Hist: OP referencia":       16,
    "Hist: Data analise":        14,
}

WRAP_COLS = {"Descricao de Item", "Descricao longa do item", "Local de Entrega"}


def _format_sheet(ws, df, title):
    ws.insert_rows(1, 2)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
    tc = ws.cell(1, 1, title)
    tc.font = Font(name="Arial", bold=True, color="FFFFFF", size=13)
    tc.fill = PatternFill("solid", fgColor="1237FF")
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    for ci, col in enumerate(df.columns, 1):
        c = ws.cell(2, ci, COL_LABELS.get(col, col))
        c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = CENTER; c.border = BORDER
    ws.row_dimensions[2].height = 22
    has_wrap = any(col in WRAP_COLS for col in df.columns)
    HIST_COL = "Hist: Ultima analise"
    GANHOU_FILL = PatternFill("solid", fgColor="C8E6C9")  # verde
    PERDEU_FILL = PatternFill("solid", fgColor="FFCDD2")  # vermelho
    SEM_FILL    = PatternFill("solid", fgColor="E8EAF6")  # cinza azulado

    for ri, (_, row) in enumerate(df.iterrows(), 3):
        is_rec = str(row.get("Recorrente", "")).strip().lower() == "sim"
        fill = REC_FILL if is_rec else (ODD_FILL if ri % 2 else EVEN_FILL)
        resultado_hist = str(row.get(HIST_COL, "")).strip().lower()

        for ci, col in enumerate(df.columns, 1):
            val = row[col]
            val = "" if pd.isna(val) else val
            c = ws.cell(ri, ci, val)
            c.font = BODY_FONT
            c.border = BORDER
            c.alignment = LEFT if col in WRAP_COLS else CENTER

            # Colunas de historico recebem cor propria
            if col in ("Hist: Ultima analise", "Hist: Dif ultima", "Hist: Dif media hist"):
                if resultado_hist == "ganhamos":
                    c.fill = GANHOU_FILL
                    c.font = Font(name="Arial", size=10, bold=(col == "Hist: Ultima analise"), color="1B5E20")
                elif resultado_hist == "perdemos":
                    c.fill = PERDEU_FILL
                    c.font = Font(name="Arial", size=10, bold=(col == "Hist: Ultima analise"), color="B71C1C")
                else:
                    c.fill = SEM_FILL
                    c.font = Font(name="Arial", size=10, color="546E7A")
            else:
                c.fill = fill
        ws.row_dimensions[ri].height = 40 if has_wrap else 18
    for ci, col in enumerate(df.columns, 1):
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(col, 18)
    ws.freeze_panes = "A3"


def _format_resumo(ws, resumo_df):
    ws.insert_rows(1, 2)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    tc = ws.cell(1, 1, "Resumo de Distribuicao")
    tc.font = Font(name="Arial", bold=True, color="FFFFFF", size=13)
    tc.fill = PatternFill("solid", fgColor="1237FF")
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    for ci, label in enumerate(["Responsavel", "Itens", "%"], 1):
        c = ws.cell(2, ci, label)
        c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = CENTER; c.border = BORDER
    ws.row_dimensions[2].height = 22
    for ri, (_, row) in enumerate(resumo_df.iterrows(), 3):
        fill = ODD_FILL if ri % 2 else EVEN_FILL
        for ci, val in enumerate([row["Responsavel"], row["Itens"], str(row["%"]) + "%"], 1):
            c = ws.cell(ri, ci, val)
            c.font = BODY_FONT; c.fill = fill; c.border = BORDER; c.alignment = CENTER
        ws.row_dimensions[ri].height = 18
    for col, w in zip("ABC", [20, 10, 10]):
        ws.column_dimensions[col].width = w


def to_excel(df, team):
    names = [m["name"] for m in team]
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        if df.empty:
            pd.DataFrame({"Aviso": ["Nenhum item encontrado."]}).to_excel(writer, sheet_name="Consolidado", index=False)
            return buf.getvalue()
        for member in team:
            sub = df[df["Responsavel"] == member["name"]]
            if not sub.empty:
                sub.to_excel(writer, sheet_name=member["name"][:31], index=False)
        df.to_excel(writer, sheet_name="Consolidado", index=False)
        resumo = pd.DataFrame([{
            "Responsavel": m["name"],
            "Itens": int((df["Responsavel"] == m["name"]).sum()),
            "%": round((df["Responsavel"] == m["name"]).sum() / len(df) * 100, 1),
        } for m in team])
        resumo.to_excel(writer, sheet_name="Resumo", index=False)

    wb = load_workbook(BytesIO(buf.getvalue()))
    for member in team:
        name = member["name"][:31]
        if name in wb.sheetnames:
            sub = df[df["Responsavel"] == member["name"]]
            _format_sheet(wb[name], sub.reset_index(drop=True), "Next Supply - " + member["name"])
    if "Consolidado" in wb.sheetnames:
        _format_sheet(wb["Consolidado"], df.reset_index(drop=True), "Next Supply - Consolidado")
    if "Resumo" in wb.sheetnames:
        resumo = pd.DataFrame([{
            "Responsavel": m["name"],
            "Itens": int((df["Responsavel"] == m["name"]).sum()),
            "%": round((df["Responsavel"] == m["name"]).sum() / len(df) * 100, 1),
        } for m in team])
        _format_resumo(wb["Resumo"], resumo)

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


# ── Interface ─────────────────────────────────────────────────────────────────

def render_topbar():
    st.markdown(
        '<div class="topbar"><div class="brand-wrap">'
        '<div class="brand-mark">N</div>'
        '<div><p class="brand-title">NEXT SUPPLY</p>'
        '<p class="brand-subtitle">Separador de OPS Petronect</p></div></div></div>',
        unsafe_allow_html=True
    )


def render_hero(df):
    total  = len(df) if df is not None and not df.empty else 0
    ops    = df["Numero da Oportunidade"].nunique() if df is not None and not df.empty else 0
    resp   = df["Responsavel"].nunique() if df is not None and not df.empty else 0
    status = "Processado" if total > 0 else "Aguardando upload"
    st.markdown(
        '<div class="hero"><div class="hero-title">Separador de OPS</div>'
        '<div class="hero-sub">Faca upload do ZIP exportado do Petronect. '
        'Para continuar o mes, carregue tambem o Excel mensal salvo anteriormente.</div>'
        '<div class="metrics">'
        '<div class="mcard"><div class="mcard-label">Oportunidades</div><div class="mcard-val accent">' + str(ops) + '</div></div>'
        '<div class="mcard"><div class="mcard-label">Itens extraidos</div><div class="mcard-val">' + str(total) + '</div></div>'
        '<div class="mcard"><div class="mcard-label">Responsaveis ativos</div><div class="mcard-val">' + str(resp) + '</div></div>'
        '<div class="mcard"><div class="mcard-label">Status</div><div class="mcard-val">' + status + '</div></div>'
        '</div></div>',
        unsafe_allow_html=True
    )


def sec(title):
    st.markdown('<div class="sec-title">' + title + '</div>', unsafe_allow_html=True)


def hr():
    st.markdown('<hr class="hr">', unsafe_allow_html=True)


# ── Main ──────────────────────────────────────────────────────────────────────

if "history" not in st.session_state:
    st.session_state["history"] = pd.DataFrame(columns=COLS)
if "last_upload" not in st.session_state:
    st.session_state["last_upload"] = None
if "df_preco" not in st.session_state:
    st.session_state["df_preco"] = pd.DataFrame()

team = sidebar_team()
render_topbar()

sec("Carregar historico mensal anterior (opcional)")
st.caption("Para continuar acumulando o mes e ativar deteccao de recorrencia, faca upload do Excel mensal salvo anteriormente.")

base_file = st.file_uploader("Excel mensal anterior (.xlsx)", type=["xlsx"], key="base_upload", label_visibility="collapsed")
if base_file:
    try:
        base_df = pd.read_excel(base_file, sheet_name="Consolidado", dtype=str)
        base_df.columns = [c.strip() for c in base_df.columns]
        inv_labels = {v: k for k, v in COL_LABELS.items()}
        base_df = base_df.rename(columns=inv_labels)
        for col in COLS:
            if col not in base_df.columns:
                base_df[col] = ""
        base_df = base_df[[c for c in COLS if c in base_df.columns]]
        for col in COLS:
            if col not in base_df.columns:
                base_df[col] = ""
        base_df = base_df[COLS]
        st.session_state["history"] = base_df.drop_duplicates(subset=["Numero da Oportunidade", "Item"], keep="last")
        st.success("Historico carregado: " + str(len(st.session_state["history"])) + " itens.")
    except Exception as e:
        st.error("Erro ao carregar base: " + str(e))

hr()

sec("Planilha de Analise de Precos (opcional)")
st.caption("Carregue a planilha de analise para ativar o historico de ganho/perda por Part Number nas colunas do cotador.")
preco_file = st.file_uploader("Analise de Precos (.xlsx)", type=["xlsx"], key="preco_upload", label_visibility="collapsed")
if preco_file:
    try:
        df_p = load_historico(preco_file.read())
        if df_p.empty:
            st.warning("Nao foi possivel ler a aba DADOS da planilha.")
        else:
            st.session_state["df_preco"] = df_p
            n_pns = df_p["_pn"].nunique()
            n_ganhos = int((df_p.get("Resultado Esperado", pd.Series()).str.strip().str.lower() == "ganhamos").sum())
            n_perdidos = int((df_p.get("Resultado Esperado", pd.Series()).str.strip().str.lower() == "perdemos").sum())
            st.success(f"Historico carregado: {len(df_p)} linhas | {n_pns} PNs unicos | {n_ganhos} ganhos | {n_perdidos} perdidos.")
    except Exception as e:
        st.error("Erro ao carregar planilha: " + str(e))

hr()

sec("Upload do arquivo ZIP")
st.caption("Selecione o arquivo compactado com os PDFs das oportunidades do dia.")
uploaded = st.file_uploader("Arquivo ZIP", type=["zip"], label_visibility="collapsed")
df_today = None

if uploaded and uploaded.name != st.session_state["last_upload"]:
    st.session_state["last_upload"] = uploaded.name
    with st.spinner("Processando PDFs..."):
        df_today, log = process_zip(
            uploaded.read(), team,
            st.session_state["history"],
            st.session_state["df_preco"] if not st.session_state["df_preco"].empty else None
        )
    if df_today is not None and not df_today.empty:
        st.session_state["history"] = (
            pd.concat([st.session_state["history"], df_today], ignore_index=True)
            .drop_duplicates(subset=["Numero da Oportunidade", "Item"], keep="last")
        )
        n_rec = int((df_today["Recorrente"] == "Sim").sum())
        msg = "OK: " + str(len(df_today)) + " item(ns) de " + str(df_today["Numero da Oportunidade"].nunique()) + " oportunidade(s)."
        if n_rec > 0:
            msg += " | " + str(n_rec) + " item(ns) recorrente(s)."
        st.success(msg)
    else:
        st.warning("Nenhum item encontrado. Confira o log abaixo.")
    with st.expander("Log de processamento"):
        for line in log:
            st.write(line)

# Compatibilidade com historicos antigos
if not st.session_state["history"].empty:
    for col, default in [("Categoria", None), ("Recorrente", "Nao")]:
        if col not in st.session_state["history"].columns:
            if col == "Categoria":
                st.session_state["history"]["Categoria"] = st.session_state["history"].get(
                    "Descricao longa do item", pd.Series()
                ).apply(categorize)
            else:
                st.session_state["history"][col] = default

df_view = st.session_state["history"] if not st.session_state["history"].empty else None
render_hero(df_view)

if df_view is not None and not df_view.empty:
    hr()
    render_alertas(df_view)
    hr()

    sec("Filtros")
    c1, c2, c3, c4 = st.columns([2, 2, 2, 3])
    with c1:
        f_resp = st.multiselect("Responsavel", options=sorted(df_view["Responsavel"].dropna().unique()))
    with c2:
        cats = sorted(df_view["Categoria"].dropna().unique()) if "Categoria" in df_view.columns else []
        f_cat = st.multiselect("Categoria", options=cats)
    with c3:
        f_rec = st.selectbox("Recorrente", ["Todos", "Sim", "Nao"])
    with c4:
        f_text = st.text_input("Busca por descricao ou numero", "")

    df_filtered = df_view.copy()
    if f_resp:
        df_filtered = df_filtered[df_filtered["Responsavel"].isin(f_resp)]
    if f_cat:
        df_filtered = df_filtered[df_filtered["Categoria"].isin(f_cat)]
    if f_rec != "Todos":
        df_filtered = df_filtered[df_filtered["Recorrente"] == f_rec]
    if f_text:
        mask = (
            df_filtered["Descricao longa do item"].str.contains(f_text, case=False, na=False) |
            df_filtered["Numero da Oportunidade"].astype(str).str.contains(f_text, na=False)
        )
        df_filtered = df_filtered[mask]

    sec("Consolidado - " + str(len(df_filtered)) + " item(ns)")
    st.dataframe(df_filtered, use_container_width=True, height=380)
    hr()

    sec("Distribuicao")
    g1, g2, g3 = st.columns(3)
    with g1:
        st.caption("Por responsavel")
        dist_r = df_view["Responsavel"].value_counts().reset_index()
        dist_r.columns = ["Responsavel", "Itens"]
        st.bar_chart(dist_r.set_index("Responsavel"), color="#3F5DFF")
    with g2:
        st.caption("Por categoria")
        dist_c = df_view["Categoria"].value_counts().head(8).reset_index()
        dist_c.columns = ["Categoria", "Itens"]
        st.bar_chart(dist_c.set_index("Categoria"), color="#1237FF")
    with g3:
        st.caption("Por tipo de oportunidade")
        dist_t = df_view["Tipo de Oportunidade"].value_counts().head(8).reset_index()
        dist_t.columns = ["Tipo", "Itens"]
        st.bar_chart(dist_t.set_index("Tipo"), color="#7B9FFF")

    hr()
    sec("Exportar")
    e1, e2, e3 = st.columns(3)
    with e1:
        src = df_today if df_today is not None and not df_today.empty else df_view
        st.download_button(
            "Excel do dia",
            data=to_excel(src, team),
            file_name="nextsupply_ops_dia.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        st.caption("Somente o ultimo upload.")
    with e2:
        st.download_button(
            "Excel mensal (acumulado)",
            data=to_excel(st.session_state["history"], team),
            file_name="nextsupply_ops_mensal.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        st.caption(str(len(st.session_state["history"])) + " itens acumulados. Salve para usar amanha.")
    with e3:
        if st.button("Limpar historico"):
            st.session_state["history"] = pd.DataFrame(columns=COLS)
            st.session_state["last_upload"] = None
            st.rerun()
        st.caption("Zera o acumulado para novo mes.")
